from __future__ import annotations

import io
import math
import os
import re
import sys
from datetime import datetime
from dataclasses import asdict, dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image
import streamlit as st

# ============================================================
# 1. Streamlit Page Config
# ============================================================
st.set_page_config(
    page_title="WIN-SQUARE | Requirement Sheet Engine",
    layout="wide",
    page_icon="🪟",
    initial_sidebar_state="expanded"
)

# ============================================================
# 2. FIX CSS: Header चालू ठेवून Sidebar Toggle Button Visible ठेवणे
# ============================================================
st.markdown("""
    <style>
    header[data-testid="stHeader"] {
        z-index: 99999 !important;
        background: transparent !important;
    }

    button[data-testid="stSidebarCollapsedControl"],
    button[data-testid="stSidebarNavCollapseButton"] {
        display: flex !important;
        visibility: visible !important;
        opacity: 1 !important;
        background-color: #FF4B4B !important;
        color: white !important;
        border-radius: 8px !important;
        position: fixed !important;
        top: 12px !important;
        left: 12px !important;
        z-index: 999999 !important;
        box-shadow: 0px 3px 8px rgba(0,0,0,0.3) !important;
    }

    button[data-testid="stSidebarCollapsedControl"] svg,
    button[data-testid="stSidebarNavCollapseButton"] svg {
        fill: white !important;
        color: white !important;
        width: 22px !important;
        height: 22px !important;
    }

    [data-testid="stStatusWidget"],
    #MainMenu, 
    footer {
        display: none !important;
        visibility: hidden !important;
    }
    </style>
""", unsafe_allow_html=True)

# State Management
if "uploader_key" not in st.session_state:
    st.session_state["uploader_key"] = 0

# ============================================================
# 3. UI Layout & Fonts CSS
# ============================================================
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
        background-color: #f4f6f9;
        color: #334155;
    }

    .main .block-container {
        padding-top: 1.5rem;
        padding-bottom: 2rem;
        max-width: 98%;
    }

    [data-testid="stSidebar"] {
        background-color: #f1f5f9;
        border-right: 1px solid #e2e8f0;
    }
    
    .quick-guide-title {
        font-size: 15px;
        font-weight: 700;
        color: #0f172a;
        margin-top: 15px;
        margin-bottom: 12px;
    }
    
    .quick-guide-step {
        font-size: 13px;
        color: #475569;
        margin-bottom: 10px;
        line-height: 1.4;
    }

    .hero-container {
        background: #ffffff;
        border-radius: 16px;
        padding: 24px 30px;
        border: 1px solid #e2e8f0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.02);
        margin-bottom: 24px;
        display: flex;
        justify-content: space-between;
        align-items: center;
    }

    .hero-title-text {
        font-size: 22px;
        font-weight: 800;
        color: #0f172a;
        margin: 0;
    }

    .hero-sub-text {
        font-size: 13px;
        color: #64748b;
        margin-top: 4px;
    }

    .step-title {
        font-size: 16px;
        font-weight: 700;
        color: #1e293b;
        margin-bottom: 12px;
        display: flex;
        align-items: center;
        gap: 8px;
    }

    .kpi-card-box {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 10px;
        padding: 16px 20px;
        box-shadow: 0 1px 2px rgba(0,0,0,0.03);
    }

    .kpi-title-lbl {
        font-size: 11px;
        font-weight: 700;
        color: #64748b;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    .kpi-val-lbl {
        font-size: 24px;
        font-weight: 800;
        color: #0f172a;
        margin-top: 6px;
    }

    /* PRIMARY BLUE BUTTON */
    div.stButton > button[kind="primary"] {
        background-color: #2563eb !important;
        background: #2563eb !important;
        border: 1px solid #2563eb !important;
        color: #ffffff !important;
        border-radius: 6px !important;
        height: 38px !important;
        padding: 0 16px !important;
        box-shadow: 0 1px 2px rgba(37, 99, 235, 0.2) !important;
    }
    div.stButton > button[kind="primary"]:hover {
        background-color: #1d4ed8 !important;
        background: #1d4ed8 !important;
    }

    /* SECONDARY RED BUTTON */
    div.stButton > button[kind="secondary"] {
        background-color: #dc2626 !important;
        background: #dc2626 !important;
        border: 1px solid #dc2626 !important;
        color: #ffffff !important;
        border-radius: 6px !important;
        height: 38px !important;
        padding: 0 16px !important;
        box-shadow: 0 1px 2px rgba(220, 38, 38, 0.2) !important;
    }
    div.stButton > button[kind="secondary"]:hover {
        background-color: #b91c1c !important;
        background: #b91c1c !important;
    }

    /* DOWNLOAD GREEN BUTTON */
    div.stDownloadButton > button {
        background-color: #059669 !important;
        background: #059669 !important;
        border: 1px solid #059669 !important;
        color: #ffffff !important;
        border-radius: 6px !important;
        height: 38px !important;
        padding: 0 16px !important;
        box-shadow: 0 1px 2px rgba(5, 150, 105, 0.2) !important;
    }
    div.stDownloadButton > button:hover {
        background-color: #047857 !important;
        background: #047857 !important;
    }

    div.stButton > button p, div.stButton > button span,
    div.stDownloadButton > button p, div.stDownloadButton > button span {
        color: #ffffff !important;
        font-weight: 500 !important;
        font-size: 13px !important;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 16px;
        border-bottom: 1px solid #e2e8f0;
    }

    .stTabs [data-baseweb="tab"] {
        height: 40px;
        white-space: pre;
        font-size: 13px;
        font-weight: 600;
        color: #64748b;
    }

    .stTabs [aria-selected="true"] {
        color: #2563eb !important;
        border-bottom: 2px solid #2563eb !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def get_image_path(filename):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.abspath("."), filename)


# =========================================================
# SIDEBAR
# =========================================================
with st.sidebar:
    logo_file = get_image_path("logo.png")
    if os.path.exists(logo_file):
        col_s1, col_s2, col_s3 = st.columns([1, 2, 1])
        with col_s2:
            st.image(Image.open(logo_file), width=110)
    else:
        st.markdown("<h2 style='text-align: center; color:#1e293b;'><b>win square</b></h2>", unsafe_allow_html=True)
    
    st.markdown("---")
    st.markdown("<div class='quick-guide-title'>💡 Quick Guide</div>", unsafe_allow_html=True)
    st.markdown(
        """
        <div class='quick-guide-step'><b>1.</b> Upload multi-sheet Excel BOQ files.</div>
        <div class='quick-guide-step'><b>2.</b> Click on <b>Merge & Process Files</b>.</div>
        <div class='quick-guide-step'><b>3.</b> Review merged glass records.</div>
        <div class='quick-guide-step'><b>4.</b> Click <b>Generate Requirement Sheet (MEASUREMENTS)</b>.</div>
        <div class='quick-guide-step'><b>5.</b> Download styled Excel with auto formulas & OC breakdown.</div>
        <div class='quick-guide-step'><b>6.</b> Use <b>Reset Data</b> to clear current workspace.</div>
        """,
        unsafe_allow_html=True
    )


# =========================================================
# HEADER HERO BANNER
# =========================================================
st.markdown(
    """
    <div class="hero-container">
        <div>
            <div class="hero-title-text">Requirement Sheet Engine</div>
            <div class="hero-sub-text">Enterprise BOQ Extraction, File Merger & Automatic Measurement Generator</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ============================================================
# Global Engine Constants & Parsing Logic
# ============================================================

HEADER_SCAN_LIMIT = 200
HEADER_REMOVE_PATTERN = r"[^A-Z0-9]"

KEYWORDS = {
    "CODE": ["CODE"],
    "WIDTH": [
        ["GL", "W"], ["GLS", "W"], ["GLASS", "W"],
        ["GL", "WIDTH"], ["GLS", "WIDTH"], ["GLASS", "WIDTH"],
        ["S", "GLS", "W"], ["SGLS", "W"], ["S", "GL", "W"],
    ],
    "HEIGHT": [
        ["GL", "H"], ["GLS", "H"], ["GLASS", "H"],
        ["GL", "HEIGHT"], ["GLS", "HEIGHT"], ["GLASS", "HEIGHT"],
        ["S", "GLS", "H"], ["SGLS", "H"], ["S", "GL", "H"],
    ],
    "QTY": ["QTY"],
    "GLASS": [["GLASS"], ["DESP"]],
}


def standardize_glass_spec(val: str) -> str:
    if pd.isna(val) or not str(val).strip():
        return "NOT SPECIFIED"
    text = str(val).strip()
    if text.lower() == "nan" or not text:
        return "NOT SPECIFIED"
    return re.sub(r"\s+", " ", text).strip()


@dataclass(slots=True)
class HeaderInfo:
    row_index: int
    code_col: Optional[int] = None
    width_col: Optional[int] = None
    height_col: Optional[int] = None
    qty_col: Optional[int] = None
    glass_col: Optional[int] = None
    columns: Dict[str, Optional[int]] = field(default_factory=dict)


@dataclass(slots=True)
class HeaderBlock:
    header: HeaderInfo
    start_row: int
    end_row: int


@dataclass(slots=True)
class GlassRecord:
    WindowCode: str
    Width: int
    Height: int
    Qty: int
    GlassType: str
    SourceFile: str
    SheetName: str


def normalize_header(text: Any) -> str:
    if pd.isna(text):
        return ""
    text = str(text).upper().strip()
    return re.sub(HEADER_REMOVE_PATTERN, "", text)


def normalize_header_row(row: pd.Series) -> List[str]:
    return [normalize_header(val) for val in row.tolist()]


def contains_keywords(text: str, keyword_groups: List[Any]) -> bool:
    if not text:
        return False
    text = normalize_header(text)
    text = re.sub(r"[^A-Z0-9]", "", text)

    for group in keyword_groups:
        if isinstance(group, str):
            group = [group]
        matched = True
        for keyword in group:
            key = re.sub(r"[^A-Z0-9]", "", normalize_header(keyword))
            if key not in text:
                matched = False
                break
        if matched:
            return True
    return False


def detect_column(header_row: List[str], keyword_groups: List[Any]) -> Optional[int]:
    for index, value in enumerate(header_row):
        text = normalize_header(value)
        text = re.sub(r"[^A-Z0-9]", "", text)
        text = (
            text.replace("GLASS", "GL")
            .replace("GLAZING", "GL")
            .replace("SGLS", "GLS")
            .replace("SGL", "GL")
            .replace("GLS.", "GLS")
        )
        for group in keyword_groups:
            if isinstance(group, str):
                group = [group]
            matched = True
            for keyword in group:
                key = re.sub(r"[^A-Z0-9]", "", normalize_header(keyword))
                if key not in text:
                    matched = False
                    break
            if matched:
                return index
    return None


def detect_header_columns(header_row: pd.Series) -> Dict[str, Optional[int]]:
    normalized = normalize_header_row(header_row)
    columns = {
        "code": detect_column(normalized, KEYWORDS["CODE"]),
        "width": detect_column(normalized, KEYWORDS["WIDTH"]),
        "height": detect_column(normalized, KEYWORDS["HEIGHT"]),
        "qty": detect_column(normalized, KEYWORDS["QTY"]),
        "glass": detect_column(normalized, KEYWORDS["GLASS"]),
    }

    if columns["width"] is None:
        for kw in [
            ["S", "GLS", "W"], ["S", "GL", "W"], ["GLS", "W"], ["GL", "W"],
            ["S", "GZ", "W"], ["GZ", "W"], ["FWIDTH"], ["F", "WIDTH"],
            ["SWIDTH"], ["S", "WIDTH"],
        ]:
            col = detect_column(normalized, kw)
            if col is not None:
                columns["width"] = col
                break

    if columns["height"] is None:
        for kw in [
            ["S", "GLS", "H"], ["S", "GL", "H"], ["GLS", "H"], ["GL", "H"],
            ["S", "GZ", "H"], ["GZ", "H"], ["FHEIGHT"], ["F", "HEIGHT"],
            ["SHEIGHT"], ["S", "HEIGHT"],
        ]:
            col = detect_column(normalized, kw)
            if col is not None:
                columns["height"] = col
                break

    return columns


def is_business_header(row: pd.Series) -> bool:
    normalized = normalize_header_row(row)
    has_code = False
    has_qty = False
    has_glass = False

    for value in normalized:
        if contains_keywords(value, KEYWORDS["CODE"]):
            has_code = True
        if contains_keywords(value, KEYWORDS["QTY"]):
            has_qty = True
        if (
            contains_keywords(value, KEYWORDS["GLASS"])
            or contains_keywords(value, KEYWORDS["WIDTH"])
            or contains_keywords(value, KEYWORDS["HEIGHT"])
        ):
            has_glass = True

    return has_code and has_qty and has_glass


def find_header_blocks(dataframe: pd.DataFrame) -> List[HeaderInfo]:
    headers: List[HeaderInfo] = []
    rows = min(len(dataframe), HEADER_SCAN_LIMIT)

    for row_number in range(rows):
        row = dataframe.iloc[row_number]
        if not is_business_header(row):
            continue

        columns = detect_header_columns(row)
        header = HeaderInfo(
            row_index=row_number,
            code_col=columns["code"],
            width_col=columns["width"],
            height_col=columns["height"],
            qty_col=columns["qty"],
            glass_col=columns["glass"],
            columns=columns,
        )
        headers.append(header)

    return headers


def build_header_blocks(dataframe: pd.DataFrame, headers: List[HeaderInfo]) -> List[HeaderBlock]:
    blocks: List[HeaderBlock] = []
    if not headers:
        return blocks

    headers = sorted(headers, key=lambda h: h.row_index)
    for i, header in enumerate(headers):
        start = header.row_index + 1
        end = (
            len(dataframe) - 1
            if i == len(headers) - 1
            else headers[i + 1].row_index - 1
        )
        blocks.append(HeaderBlock(header=header, start_row=start, end_row=end))

    return blocks


def score_business_sheet(df: pd.DataFrame) -> int:
    score = 0
    scan_rows = min(len(df), HEADER_SCAN_LIMIT)

    for r in range(scan_rows):
        row = normalize_header_row(df.iloc[r])
        text = " ".join(row)
        if "CODE" in text:
            score += 5
        if "QTY" in text:
            score += 8
        if "GLASS" in text:
            score += 8
        if re.search(r"\bGL.*W\b", text):
            score += 10
        if re.search(r"\bGL.*H\b", text):
            score += 10

    return score


def find_business_sheets(workbook: Dict[str, pd.DataFrame]) -> List[Tuple[str, pd.DataFrame]]:
    business_sheets = []
    threshold = 35

    for sheet_name, df in workbook.items():
        score = score_business_sheet(df)
        accept_sheet = False

        if score >= threshold:
            accept_sheet = True
        else:
            header_blocks = find_header_blocks(df)
            if header_blocks:
                accept_sheet = True

        if accept_sheet:
            business_sheets.append((sheet_name, df))

    return business_sheets


def safe_numeric(value: Any) -> Optional[int]:
    if pd.isna(value):
        return None
    try:
        val = float(value)
        if val <= 0:
            return None
        return int(math.floor(val + 0.5))
    except Exception:
        return None


def build_window_code(row: pd.Series, header: HeaderInfo) -> Optional[str]:
    if header.code_col is None:
        return None

    val = row.iloc[header.code_col]
    if pd.isna(val):
        return None

    code = str(val).strip()
    if code == "" or code.lower() == "nan":
        return None

    code = re.sub(r"\.0$", "", code)
    code = re.sub(r"\s+", " ", code).strip()

    if header.code_col + 1 < len(row):
        next_val = row.iloc[header.code_col + 1]
        if not pd.isna(next_val):
            next_str = str(next_val).strip()
            next_str = re.sub(r"\.0$", "", next_str)

            if next_str and next_str.lower() != "nan":
                if code.isdigit():
                    code = f"{code} {next_str}"
                elif not re.search(r"\b[A-Z]*\d+\b", code, re.I):
                    if re.match(r"^(W|D|K|CW|NW)\d*$", next_str, re.I):
                        code = f"{code} {next_str}"

    return re.sub(r"\s+", " ", code).strip()


def is_record_start(row: pd.Series, header: HeaderInfo) -> bool:
    try:
        code_parts = []
        for col in range(min(3, len(row))):
            value = row.iloc[col]
            if pd.isna(value):
                continue
            val_str = str(value).strip()
            if val_str == "":
                continue
            code_parts.append(normalize_header(val_str))

        if not code_parts:
            return False
        code = " ".join(code_parts).strip()

        invalid = ("CODE", "FWIDTH", "FHEIGHT", "GLSW", "GLSH", "GLASS", "QTY", "DESCRIPTION")
        if code in invalid or any(inv == code for inv in invalid):
            return False

        upper = code.upper()
        if any(w in upper for w in ["FRAME", "WINDOW", "SECTION", "PROFILE"]):
            return False
        if not re.search(r"[A-Z0-9]", upper):
            return False

        return True
    except Exception:
        return False


def build_record_buffers(dataframe: pd.DataFrame, block: HeaderBlock) -> List[pd.DataFrame]:
    buffers: List[pd.DataFrame] = []
    current_rows = []

    for row_no in range(block.start_row, block.end_row + 1):
        row = dataframe.iloc[row_no]
        if is_record_start(row, block.header):
            if current_rows:
                buffers.append(pd.DataFrame(current_rows))
                current_rows = []
        current_rows.append(row)

    if current_rows:
        buffers.append(pd.DataFrame(current_rows))

    return buffers


def collect_numeric_from_buffer(buffer: pd.DataFrame, column_index: Optional[int]) -> Optional[int]:
    if column_index is None:
        return None
    for _, row in buffer.iterrows():
        if column_index >= len(row):
            continue
        val = safe_numeric(row.iloc[column_index])
        if val is not None:
            return val
    return None


def collect_glass(buffer: pd.DataFrame, header: HeaderInfo) -> Optional[str]:
    parts = []
    seen = set()

    if header.glass_col is not None:
        for _, row in buffer.iterrows():
            if header.glass_col >= len(row):
                continue
            val = row.iloc[header.glass_col]
            if pd.isna(val):
                continue
            val_str = re.sub(r"\s+", " ", str(val).strip()).strip()
            if val_str == "" or val_str.upper() in seen:
                continue
            seen.add(val_str.upper())
            parts.append(val_str)

    if parts:
        return " ".join(parts)

    for _, row in buffer.iterrows():
        for cell in row.values:
            if pd.isna(cell):
                continue
            cell_str = str(cell).strip()
            upper_cell = cell_str.upper()
            if any(k in upper_cell for k in ["LAMINATED", "TOUGHENED", "DGU", "SGU", "CLEAR", "FROSTED", "MM"]):
                if not any(k in upper_cell for k in ["CODE", "HANDLE", "LOCK", "COLOR", "FRAME", "WIDTH", "HEIGHT"]):
                    return cell_str

    return None


def parse_header_block(dataframe: pd.DataFrame, block: HeaderBlock, source_file: str, sheet_name: str) -> List[GlassRecord]:
    records: List[GlassRecord] = []
    buffers = build_record_buffers(dataframe, block)

    for buffer in buffers:
        first_row = buffer.iloc[0]
        window = build_window_code(first_row, block.header)

        width = collect_numeric_from_buffer(buffer, block.header.width_col)
        height = collect_numeric_from_buffer(buffer, block.header.height_col)
        qty = collect_numeric_from_buffer(buffer, block.header.qty_col)
        glass_raw = collect_glass(buffer, block.header)

        if not window or width is None or height is None:
            continue
        if qty is None:
            qty = 1

        if glass_raw and "FROSTED" in str(glass_raw).upper():
            continue

        glass = standardize_glass_spec(glass_raw)

        records.append(
            GlassRecord(
                WindowCode=window,
                Width=width,
                Height=height,
                Qty=qty,
                GlassType=glass,
                SourceFile=source_file,
                SheetName=sheet_name,
            )
        )

    return records


def parse_business_sheet(dataframe: pd.DataFrame, source_file: str, sheet_name: str) -> List[GlassRecord]:
    headers = find_header_blocks(dataframe)
    blocks = build_header_blocks(dataframe, headers)
    all_records: List[GlassRecord] = []

    for block in blocks:
        all_records.extend(parse_header_block(dataframe, block, source_file, sheet_name))

    return all_records


def load_excel_with_calculated_values(file) -> Dict[str, pd.DataFrame]:
    file_bytes = io.BytesIO(file.read())
    file.seek(0)
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    workbook_dict = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data = sheet.values
        cols = next(data, None)
        if cols is None:
            continue

        data_rows = list(data)
        if cols:
            data_rows.insert(0, cols)

        df = pd.DataFrame(data_rows)
        workbook_dict[sheet_name] = df

    return workbook_dict


# ============================================================
# PROCESS FILES
# ============================================================
def process_uploaded_files(uploaded_files) -> pd.DataFrame:
    all_records = []

    for file in uploaded_files:
        try:
            workbook_dict = load_excel_with_calculated_values(file)
            filtered_workbook = {}
            for s_name, df in workbook_dict.items():
                s_name_upper = s_name.upper().strip()
                if any(k in s_name_upper for k in ["CLIENT", "DETAIL", "QUOTE"]):
                    continue
                filtered_workbook[s_name] = df

            business_sheets = find_business_sheets(filtered_workbook)
            for sheet_name, df in business_sheets:
                records = parse_business_sheet(df, file.name, sheet_name)
                all_records.extend(records)

        except Exception as e:
            st.error(f"Error processing file {file.name}: {e}")

    return pd.DataFrame([asdict(r) for r in all_records]).reset_index(drop=True) if all_records else pd.DataFrame()


# ============================================================
# STEP 1: FILE UPLOAD SECTION
# ============================================================
st.markdown("<div class='step-title'>📁 Step 1: Upload BOQ Excel Files</div>", unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    "Upload BOQ Excel Files",
    type=["xlsx", "xlsm", "xls"],
    accept_multiple_files=True,
    label_visibility="collapsed",
    key=f"boq_uploader_{st.session_state['uploader_key']}"
)

st.markdown("<br>", unsafe_allow_html=True)

btn_col1, btn_col2, _ = st.columns([1, 1, 6])

with btn_col1:
    if st.button("Merge & Process Files", type="primary", use_container_width=False):
        if uploaded_files:
            with st.spinner("Extracting & Merging Records..."):
                df_merged = process_uploaded_files(uploaded_files)
                if not df_merged.empty:
                    st.session_state["merged_df"] = df_merged
                    st.toast(f"Successfully Extracted {len(df_merged)} Non-Frosted Glass Records!", icon="✅")
                else:
                    st.error("⚠️ No valid non-frosted glass records found.")
        else:
            st.warning("Please upload Excel file(s) first!")

with btn_col2:
    if st.button("Reset Data", type="secondary", use_container_width=False):
        for key in ["merged_df", "req_df_preview", "req_bytes", "req_generated"]:
            if key in st.session_state:
                del st.session_state[key]
        st.session_state["uploader_key"] += 1
        st.rerun()


# ============================================================
# EXTRACTED MASTER GLASS RECORDS TABLE
# ============================================================
if "merged_df" in st.session_state:
    df_merged = st.session_state["merged_df"]

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='step-title'>📋 Extracted Glass Records (Frosted Excluded)</div>", unsafe_allow_html=True)

    f_col1, f_col2 = st.columns([2, 2])
    with f_col1:
        search_query = st.text_input("🔍 Quick Search (Window Code / Glass Spec)", placeholder="Type to filter...")
    with f_col2:
        glass_types = ["ALL"] + sorted(list(df_merged["GlassType"].unique()))
        selected_glass = st.selectbox("Filter by Glass Spec", glass_types)

    filtered_df = df_merged.copy()
    if search_query:
        filtered_df = filtered_df[
            filtered_df["WindowCode"].str.contains(search_query, case=False, na=False) |
            filtered_df["GlassType"].str.contains(search_query, case=False, na=False)
        ]
    if selected_glass != "ALL":
        filtered_df = filtered_df[filtered_df["GlassType"] == selected_glass]

    filtered_display_df = filtered_df.copy()
    if "Sr. No." not in filtered_display_df.columns:
        filtered_display_df.insert(0, "Sr. No.", range(1, len(filtered_display_df) + 1))

    st.dataframe(filtered_display_df, use_container_width=True, height=280, hide_index=True)
    st.caption(f"Showing {len(filtered_df)} of {len(df_merged)} extracted non-frosted records")

    # ============================================================
    # STEP 2: REQUIREMENT GENERATION & KPI CARDS
    # ============================================================
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='step-title'>⚡ Step 2: Generate Official Requirement Sheet</div>", unsafe_allow_html=True)

    if st.button("⚡ GENERATE REQUIREMENT SHEET (MEASUREMENTS)", type="primary", use_container_width=False):
        with st.spinner("Calculating SQFT and generating Excel sheet..."):
            df_req_preview = df_merged.copy()
            df_req_preview["SQFT"] = ((df_req_preview["Width"] * df_req_preview["Height"]) / 92903.04).round(6)
            df_req_preview["TTL SQFT"] = (df_req_preview["SQFT"] * df_req_preview["Qty"]).round(6)

            df_req_preview.insert(0, "Sr.No", range(1, len(df_req_preview) + 1))
            df_req_preview = df_req_preview.rename(
                columns={
                    "WindowCode": "WINDOW CODE",
                    "Width": "WIDTH",
                    "Height": "HEIGHT",
                    "Qty": "QTY",
                    "GlassType": "REMARKS",
                }
            )

            preview_cols = ["Sr.No", "WINDOW CODE", "WIDTH", "HEIGHT", "SQFT", "QTY", "TTL SQFT", "REMARKS"]
            df_req_preview = df_req_preview[preview_cols]
            st.session_state["req_df_preview"] = df_req_preview

            # OpenPyXL Sheet Processing
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "MEASUREMENTS"
            ws.views.sheetView[0].showGridLines = True

            # FONT STYLES (Requirement: Data 11pt, Header 12pt)
            title_font = Font(name="Calibri", size=12, bold=True, color="000000")
            header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
            data_font = Font(name="Calibri", size=11)
            total_font = Font(name="Calibri", size=12, bold=True)

            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # HIGH-CONTRAST DISTINCT PASTEL COLORS FOR DIFFERENT GLASS TYPES
            glass_color_palette = [
                "E2EFDA",  # Light Mint Green (Row 3, Glass 1)
                "FFF2CC",  # Soft Cream Yellow (Row 4, Glass 2)
                "DDEBF7",  # Sky Blue (Glass 3)
                "FCE4D6",  # Light Peach (Glass 4)
                "E8D8F8",  # Lavender (Glass 5)
                "F8CECC",  # Soft Pink (Glass 6)
                "E1F5FE",  # Ice Blue (Glass 7)
                "FFF3E0"   # Pastel Orange (Glass 8)
            ]
            
            unique_glasses = list(df_merged["GlassType"].unique())
            glass_fill_map = {}
            for g_idx, g_spec in enumerate(unique_glasses):
                c_hex = glass_color_palette[g_idx % len(glass_color_palette)]
                glass_fill_map[g_spec] = PatternFill(start_color=c_hex, end_color=c_hex, fill_type="solid")

            thin_border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9"),
            )
            thick_top_double_bottom = Border(
                top=Side(style="thin", color="000000"),
                bottom=Side(style="double", color="000000"),
            )

            # AUTO CURRENT DATE GENERATION LOGIC
            today_str = datetime.now().strftime("%d %b %Y").upper()
            title_text = f"1 WIN-SQUARE {today_str}"
            
            ws.merge_cells("A1:H1")
            cell_a1 = ws["A1"]
            cell_a1.value = title_text
            cell_a1.font = title_font
            cell_a1.alignment = Alignment(horizontal="center", vertical="center")
            
            for col_i in range(1, 9):
                c_cell = ws.cell(row=1, column=col_i)
                c_cell.fill = title_fill
                c_cell.border = thin_border

            # ROW 2: HEADERS (12 Font Size)
            headers = ["Sr.No", "WINDOW CODE", "WIDTH", "HEIGHT", "SQFT", "QTY", "TTL SQFT", "REMARKS"]
            for col_i, h_text in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_i, value=h_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            # ROW 3+: DATA ROWS (11 Font Size)
            for idx, row in df_merged.iterrows():
                r_idx = idx + 3
                sqft_formula = f"=ROUND((C{r_idx}*D{r_idx})/92903.04, 6)"
                ttl_sqft_formula = f"=E{r_idx}*F{r_idx}"

                row_data = [
                    idx + 1, row["WindowCode"], row["Width"], row["Height"],
                    sqft_formula, row["Qty"], ttl_sqft_formula, row["GlassType"],
                ]

                # Distinct Row Fill according to Glass Spec
                current_fill = glass_fill_map.get(row["GlassType"], PatternFill(fill_type=None))

                for col_i, val in enumerate(row_data, 1):
                    cell = ws.cell(row=r_idx, column=col_i, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.fill = current_fill

                    if col_i in [3, 4]:
                        cell.number_format = "0"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_i == 5:
                        cell.number_format = "0.000000"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_i == 6:
                        cell.number_format = "0"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_i == 7:
                        cell.number_format = "0.000000"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_i == 1:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            # TOTAL ROW
            tot_row = len(df_merged) + 3
            ws.cell(row=tot_row, column=5, value="TOTAL").font = total_font
            ws.cell(row=tot_row, column=5).alignment = Alignment(horizontal="right", vertical="center")

            qty_sum = ws.cell(row=tot_row, column=6, value=f"=SUM(F3:F{tot_row-1})")
            qty_sum.font = total_font
            qty_sum.number_format = "0"
            qty_sum.alignment = Alignment(horizontal="center", vertical="center")

            ttl_sqft_sum = ws.cell(row=tot_row, column=7, value=f"=SUM(G3:G{tot_row-1})")
            ttl_sqft_sum.font = total_font
            ttl_sqft_sum.number_format = "0.000000"
            ttl_sqft_sum.alignment = Alignment(horizontal="right", vertical="center")

            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=tot_row, column=c)
                cell.fill = total_fill
                cell.border = thick_top_double_bottom

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

            output = io.BytesIO()
            wb.save(output)
            st.session_state["req_bytes"] = output.getvalue()
            st.session_state["req_generated"] = True

    # Render KPI Cards & Live Preview
    if st.session_state.get("req_generated"):
        st.markdown("<br>", unsafe_allow_html=True)
        
        req_df = st.session_state["req_df_preview"]
        tot_items = len(req_df)
        tot_qty = req_df["QTY"].sum()
        tot_area = req_df["TTL SQFT"].sum().round(2)

        k1, k2, k3 = st.columns(3)
        with k1:
            st.markdown(f"<div class='kpi-card-box'><div class='kpi-title-lbl'>TOTAL ITEMS</div><div class='kpi-val-lbl'>{tot_items}</div></div>", unsafe_allow_html=True)
        with k2:
            st.markdown(f"<div class='kpi-card-box'><div class='kpi-title-lbl'>TOTAL GLASS QUANTITY</div><div class='kpi-val-lbl'>{tot_qty} Pcs</div></div>", unsafe_allow_html=True)
        with k3:
            st.markdown(f"<div class='kpi-card-box'><div class='kpi-title-lbl'>TOTAL GLASS SQFT (NON-FROSTED)</div><div class='kpi-val-lbl'>{tot_area:,.2f} Sq.Ft</div></div>", unsafe_allow_html=True)
    
        st.markdown("<br>", unsafe_allow_html=True)

        tab1, tab2, tab3 = st.tabs([
            "📄 MEASUREMENTS Live Preview", 
            "📊 OC Wise Summary (Glass SQFT)", 
            "🧩 Glass Type Breakdown"
        ])

        with tab1:
            st.dataframe(req_df, use_container_width=True, height=350, hide_index=True)

        with tab2:
            df_merged_copy = df_merged.copy()
            df_merged_copy["Total_SQFT"] = ((df_merged_copy["Width"] * df_merged_copy["Height"]) / 92903.04) * df_merged_copy["Qty"]

            def clean_glass_name(text):
                text = str(text).upper().strip()
                text = re.sub(r"\s+", " ", text)
                text = text.replace("TOUGHENED", "THGN").replace("TOUGHNED", "THGN")
                return text

            df_merged_copy["CleanGlassType"] = df_merged_copy["GlassType"].apply(clean_glass_name)

            def make_glass_string(group):
                summary = group.groupby("CleanGlassType")["Qty"].sum()
                details = [
                    f"{g_type} - {qty}" 
                    for g_type, qty in summary.items() 
                    if g_type != "NOT SPECIFIED"
                ]
                return ", ".join(details) if details else "-"

            glass_details_series = (
                df_merged_copy.groupby("SourceFile")
                .apply(make_glass_string, include_groups=False)
                .reset_index(name="GLASS DETAILS")
            )

            oc_summary = (
                df_merged_copy.groupby("SourceFile", as_index=False)
                .agg(
                    Qty=("Qty", "sum"),
                    Total_SQFT=("Total_SQFT", "sum")
                )
            )

            oc_summary = pd.merge(oc_summary, glass_details_series, on="SourceFile")
            oc_summary["Total_SQFT"] = oc_summary["Total_SQFT"].round(2)
            
            oc_summary.columns = ["SourceFile (OC Name)", "Qty (Pcs)", "Total Glass SQFT", "GLASS DETAILS"]
            
            if "Sr. No." not in oc_summary.columns:
                oc_summary.insert(0, "Sr. No.", range(1, len(oc_summary) + 1))

            st.dataframe(oc_summary, use_container_width=True, hide_index=True)

        with tab3:
            df_glass_copy = df_merged.copy()

            def clean_glass_name(text):
                text = str(text).upper().strip()
                text = re.sub(r"\s+", " ", text)
                text = text.replace("TOUGHNED", "TOUGHENED")
                return text

            df_glass_copy["CleanGlassType"] = df_glass_copy["GlassType"].apply(clean_glass_name)
            df_glass_filtered = df_glass_copy[df_glass_copy["CleanGlassType"] != "NOT SPECIFIED"]

            glass_breakdown = (
                df_glass_filtered.groupby("CleanGlassType", as_index=False)["Qty"]
                .sum()
                .sort_values(by="Qty", ascending=False)
            )

            glass_breakdown.columns = ["GlassType", "Qty"]
            glass_breakdown.insert(0, "Sr. No.", range(1, len(glass_breakdown) + 1))

            st.dataframe(glass_breakdown, use_container_width=True, hide_index=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.success("✅ Requirement Excel Sheet Ready! Formatted with distinct row colors and dynamic auto-date.")
        
        st.download_button(
            label="📥 DOWNLOAD OFFICIAL REQUIREMENT SHEET (.XLSX)",
            data=st.session_state["req_bytes"],
            file_name=f"REQUIREMENT_SHEET_MEASUREMENTS_{datetime.now().strftime('%d_%b_%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False
        )
